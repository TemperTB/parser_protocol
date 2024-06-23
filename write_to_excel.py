import openpyxl

from const import ALLUM, ALPHA, AMMIAK, APAV, BARIY, BERILLIY, BETTA, BOR, BPK, BPK_POLN, BROM, CINK, CISTI, COMPANY, CVET, D_24, DATA, DATE, ECOLI, ENT, FENOL, FTOR, GELMINT, GIDROKARBONAT, HLORID, HROM, KADMIY, KLOSTRID, KOBALT, KOLIFAG, KREMNIY, LINDAN, LITIY, LYAMBLIY, MAGNIY, MARGANEC, MED, MISHYAK, MUTN, NAME, NEFT, NIKEL, NITRAT, NITRIT, NUMBER, OKB, OMCH, PER_OKISL, PH, PIT, POINT, POLIFOSFAT, RADON, RASTV_KISLOROD, RESULT, RTUT, SALMONELLA, SELEN, SHELOCH, STOK, STRONCIY, SUH_OST, SULFAT, SVINEC, TYPE, VKUS, VZV_VESH, XPK, ZAPAH_20, ZAPAH_60, ZHELEZO, ZHESTKOST, ZHIZN_GELMINT

START_ROW = 5
EXCEL_COLUMNS_NAME = 1
EXCEL_COLUMNS_COMPANY = 2
EXCEL_COLUMNS_DATA = {TYPE: 3, NUMBER: 5, DATE: 6, POINT: 4}
start_column_result = 6
EXCEL_COLUMNS_PIT_RESULTS = {
    ZAPAH_20: start_column_result + 1,
    ZAPAH_60: start_column_result + 2,
    VKUS: start_column_result + 3,
    CVET: start_column_result + 4,
    MUTN: start_column_result + 5,
    VZV_VESH: start_column_result + 6,
    PH: start_column_result + 7,
    PER_OKISL: start_column_result + 8,
    ZHESTKOST: start_column_result + 9,
    SUH_OST: start_column_result + 10,
    NEFT: start_column_result + 11,
    AMMIAK: start_column_result + 12,
    NITRIT: start_column_result + 13,
    NITRAT: start_column_result + 14,
    HLORID: start_column_result + 15,
    ZHELEZO: start_column_result + 16,
    MARGANEC: start_column_result + 17,
    MED: start_column_result + 18,
    SVINEC: start_column_result + 19,
    CINK: start_column_result + 20,
    ALLUM: start_column_result + 21,
    ECOLI: start_column_result + 22,
    OMCH: start_column_result + 23,
    OKB: start_column_result + 24,
    ENT: start_column_result + 25,
    KOLIFAG: start_column_result + 26,
    KLOSTRID: start_column_result + 27,
    SALMONELLA: start_column_result + 28,
    POLIFOSFAT: start_column_result + 29,
    APAV: start_column_result + 30,
    FENOL: start_column_result + 31,
    SULFAT: start_column_result + 32,
    RTUT: start_column_result + 33,
    BOR: start_column_result + 34,
    KREMNIY: start_column_result + 35,
    MISHYAK: start_column_result + 36,
    FTOR: start_column_result + 37,
    BROM: start_column_result + 38,
    LITIY: start_column_result + 39,
    BARIY: start_column_result + 40,
    BERILLIY: start_column_result + 41,
    SELEN: start_column_result + 42,
    STRONCIY: start_column_result + 43,
    LINDAN: start_column_result + 44,
    D_24: start_column_result + 45,
    NIKEL: start_column_result + 46,
    KADMIY: start_column_result + 47,
    GIDROKARBONAT: start_column_result + 48,
    XPK: start_column_result + 49,
    BPK: start_column_result + 50,
    GELMINT: start_column_result + 51,
    ZHIZN_GELMINT: start_column_result + 51,
    LYAMBLIY: start_column_result + 52,
    RADON: start_column_result + 53,
    ALPHA: start_column_result + 54,
    BETTA: start_column_result + 55,
    RASTV_KISLOROD: start_column_result + 56,
    KOBALT: start_column_result + 57,
    HROM: start_column_result + 58,
    MAGNIY: start_column_result + 59,
}
EXCEL_COLUMNS_STOK_RESULTS = {
    VZV_VESH: start_column_result + 1,
    PH: start_column_result + 2,
    SUH_OST: start_column_result + 3,
    NEFT: start_column_result + 4,
    AMMIAK: start_column_result + 5,
    NITRIT: start_column_result + 6,
    NITRAT: start_column_result + 7,
    HLORID: start_column_result + 8,
    ZHELEZO: start_column_result + 9,
    ECOLI: start_column_result + 10,
    ENT: start_column_result + 11,
    OKB: start_column_result + 12,
    KOLIFAG: start_column_result + 13,
    SALMONELLA: start_column_result + 14,
    POLIFOSFAT: start_column_result + 15,
    APAV: start_column_result + 16,
    SULFAT: start_column_result + 17,
    XPK: start_column_result + 18,
    BPK: start_column_result + 19,
    BPK_POLN: start_column_result + 20,
    CISTI: start_column_result + 21,
    ZHIZN_GELMINT: start_column_result + 22,
    SHELOCH: start_column_result + 23,
}

def write_to_excel(path_template, path_xlsx, protocols):
    wb = openpyxl.load_workbook(path_template + 'Template.xlsx')
    pit_water_row = START_ROW
    st_wate_row = START_ROW
    for protocol in protocols:
        if protocol[DATA][TYPE] != STOK:
            sheet = wb['Питьевая, Поверхностная']
            sheet.cell(row=pit_water_row, column=EXCEL_COLUMNS_NAME).value = protocol[NAME]
            sheet.cell(row=pit_water_row, column=EXCEL_COLUMNS_COMPANY).value = protocol[COMPANY]
            for key in EXCEL_COLUMNS_DATA.keys():
                sheet.cell(row=pit_water_row, column=EXCEL_COLUMNS_DATA[key]).value = protocol[DATA][key]
            for key in EXCEL_COLUMNS_PIT_RESULTS.keys():
                sheet.cell(row=pit_water_row, column=EXCEL_COLUMNS_PIT_RESULTS[key]).value = protocol[RESULT][key]
            pit_water_row = pit_water_row + 1
        elif protocol[DATA][TYPE] == STOK:
            sheet = wb['Сточная']
            sheet.cell(row=st_wate_row, column=EXCEL_COLUMNS_NAME).value = protocol[NAME]
            sheet.cell(row=st_wate_row, column=EXCEL_COLUMNS_COMPANY).value = protocol[COMPANY]
            for key in EXCEL_COLUMNS_DATA.keys():
                sheet.cell(row=st_wate_row, column=EXCEL_COLUMNS_DATA[key]).value = protocol[DATA][key]
            for key in EXCEL_COLUMNS_STOK_RESULTS.keys():
                sheet.cell(row=st_wate_row, column=EXCEL_COLUMNS_STOK_RESULTS[key]).value = protocol[RESULT][key]
            st_wate_row = st_wate_row + 1
    wb.save(path_xlsx + 'Results.xlsx')
    wb.close
